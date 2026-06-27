#!/usr/bin/env python3
"""
Extract Excel data and generate SQL seed file.
Run: python3 generate_seed.py > sql/seed_data.sql
"""
import openpyxl
import random
from datetime import datetime, timedelta

def esc(val):
    """Escape a value for SQL"""
    if val is None:
        return 'NULL'
    s = str(val).replace("'", "\\'").strip()
    if s == '' or s == 'None':
        return 'NULL'
    return f"'{s}'"

def num(val):
    """Convert to number or 0"""
    if val is None:
        return 0
    try:
        return float(val)
    except:
        return 0

# ============================================================
print("-- ============================================================")
print("-- Jewelry BI System - Seed Data")
print("-- Generated from Excel dummy data files")
print("-- ============================================================")
print("-- Run: ddev mysql < sql/seed_data.sql")
print("-- ============================================================")
print()
print("USE db;")
print()

# ============================================================
# 1. CATEGORIES - Extract from inventory
# ============================================================
wb = openpyxl.load_workbook('/home/ayush/jwellery/data/Dummy_Jewellery_Inventory.xlsx')
ws = wb.active

# Collect unique categories
categories = {}
cat_id = 1
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    cat = str(row[1]).strip() if row[1] else None
    if cat and cat not in categories and cat != 'None':
        categories[cat] = cat_id
        cat_id += 1

cat_descriptions = {
    'Chain': 'Gold and diamond chains',
    'Ring': 'Gold, diamond and gemstone rings',
    'Bangle': 'Gold and diamond bangles',
    'Bracelet': 'Gold and platinum bracelets',
    'Necklace Set': 'Necklace and earring sets',
    'Earring': 'Studs, drops and jhumkas',
    'Pendant': 'Gold and diamond pendants',
    'Mangalsutra': 'Traditional mangalsutra designs',
    'Nose Pin': 'Gold and diamond nose pins',
    'Anklet': 'Gold and silver anklets',
}

print("-- ============================================================")
print("-- CATEGORIES")
print("-- ============================================================")
print("INSERT INTO categories (category_id, category_name, description) VALUES")
lines = []
for cat_name, cid in sorted(categories.items(), key=lambda x: x[1]):
    desc = cat_descriptions.get(cat_name, f'{cat_name} jewelry items')
    lines.append(f"  ({cid}, {esc(cat_name)}, {esc(desc)})")
print(",\n".join(lines) + ";")
print()

# ============================================================
# 2. PRODUCTS - From inventory data
# ============================================================
print("-- ============================================================")
print("-- PRODUCTS")
print("-- ============================================================")

products = {}  # sku -> product_id
product_id = 1
product_rows = []

# Gold rates for price calculation (per gram)
gold_rates = {'18KT': 5000, '22KT': 6200, '24KT': 6800, '14KT': 3800}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    sku = str(row[0]).strip() if row[0] else None
    if not sku or sku == 'None':
        continue
    
    cat_name = str(row[1]).strip() if row[1] else 'Other'
    purity = str(row[2]).strip() if row[2] else None
    color = str(row[3]).strip() if row[3] else None
    gross_wt = num(row[4])
    net_wt = num(row[5])
    hallmark = str(row[6]).strip() if row[6] else None
    stone_wt = num(row[7])
    details = str(row[8]).strip() if row[8] else None
    
    cat_id_val = categories.get(cat_name, 1)
    
    # Generate product name
    product_name = f"{cat_name} {purity}" if purity else cat_name
    if color and 'Gold' in str(color):
        product_name = f"{color} {cat_name}"
    
    # Calculate prices based on weight and gold rate
    rate_per_gram = gold_rates.get(str(purity), 5500)
    wt = net_wt if net_wt > 0 else gross_wt
    base_price = wt * rate_per_gram
    making = round(base_price * random.uniform(0.08, 0.15), 2)
    cost_price = round(base_price + making, 2)
    selling_price = round(cost_price * random.uniform(1.15, 1.30), 2)
    
    products[sku] = product_id
    product_rows.append((
        product_id, cat_id_val, product_name, sku, 'Gold',
        purity, color, gross_wt, net_wt, stone_wt,
        hallmark, making, cost_price, selling_price, details
    ))
    product_id += 1

# Print in batches of 50
for i in range(0, len(product_rows), 50):
    batch = product_rows[i:i+50]
    print(f"INSERT INTO products (product_id, category_id, product_name, sku, metal_type, purity, color, gross_weight, net_weight, stone_weight, hallmark_code, making_charges, cost_price, selling_price, additional_info) VALUES")
    lines = []
    for p in batch:
        lines.append(f"  ({p[0]}, {p[1]}, {esc(p[2])}, {esc(p[3])}, {esc(p[4])}, {esc(p[5])}, {esc(p[6])}, {p[7]}, {p[8]}, {p[9]}, {esc(p[10])}, {p[11]}, {p[12]}, {p[13]}, {esc(p[14])})")
    print(",\n".join(lines) + ";")
    print()

# ============================================================
# 3. INVENTORY - Generate stock for each product
# ============================================================
print("-- ============================================================")
print("-- INVENTORY")
print("-- ============================================================")

random.seed(42)  # Reproducible
inv_rows = []
for sku, pid in products.items():
    qty = random.choice([0, 0, 1, 2, 3, 4, 5, 6, 7, 8, 10, 12, 15, 20, 25, 30])
    reorder = random.choice([3, 5, 5, 5, 8, 10])
    last_restock = f"2026-{random.randint(1,6):02d}-{random.randint(1,28):02d}" if qty > 0 else None
    inv_rows.append((pid, qty, reorder, last_restock))

for i in range(0, len(inv_rows), 50):
    batch = inv_rows[i:i+50]
    print("INSERT INTO inventory (product_id, quantity_in_stock, reorder_level, last_restocked) VALUES")
    lines = []
    for inv in batch:
        lines.append(f"  ({inv[0]}, {inv[1]}, {inv[2]}, {esc(inv[3])})")
    print(",\n".join(lines) + ";")
    print()

# ============================================================
# 4. CUSTOMERS - From customer report
# ============================================================
print("-- ============================================================")
print("-- CUSTOMERS")
print("-- ============================================================")

wb2 = openpyxl.load_workbook('/home/ayush/jwellery/data/Jaipur_Focused_Customer_Report.xlsx')
ws2 = wb2.active

customer_ids = {}  # customer_code -> db_id
cust_db_id = 1

cust_rows = []
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
    code = str(row[0]).strip() if row[0] else None
    if not code or code == 'None':
        continue
    
    name = str(row[1]).strip() if row[1] else 'Unknown'
    phone = str(row[2]).strip() if row[2] else None
    city = str(row[3]).strip() if row[3] else None
    state = str(row[4]).strip() if row[4] else None
    address = str(row[5]).strip() if row[5] else None
    gender = str(row[6]).strip() if row[6] else None
    since = row[7]
    
    if isinstance(since, datetime):
        since_str = since.strftime('%Y-%m-%d')
    else:
        since_str = str(since).strip() if since else None
    
    # Determine customer type based on registration date
    cust_type = 'Regular'
    if since_str:
        try:
            since_date = datetime.strptime(since_str[:10], '%Y-%m-%d')
            if since_date.year <= 2022:
                cust_type = 'VIP'
            elif since_date.year <= 2023:
                cust_type = 'Premium'
        except:
            pass
    
    # Fix gender if needed
    if gender and gender not in ('Male', 'Female', 'Other'):
        gender = None
    
    customer_ids[code] = cust_db_id
    cust_rows.append((cust_db_id, code, name, phone, city, state, address, gender, since_str, cust_type))
    cust_db_id += 1

for i in range(0, len(cust_rows), 50):
    batch = cust_rows[i:i+50]
    print("INSERT INTO customers (customer_id, customer_code, customer_name, phone, city, state, address, gender, customer_since, customer_type) VALUES")
    lines = []
    for c in batch:
        lines.append(f"  ({c[0]}, {esc(c[1])}, {esc(c[2])}, {esc(c[3])}, {esc(c[4])}, {esc(c[5])}, {esc(c[6])}, {esc(c[7])}, {esc(c[8])}, {esc(c[9])})")
    print(",\n".join(lines) + ";")
    print()

# ============================================================
# 5. SALES + SALE_ITEMS - From sales report
# ============================================================
print("-- ============================================================")
print("-- SALES AND SALE_ITEMS")
print("-- ============================================================")

wb3 = openpyxl.load_workbook('/home/ayush/jwellery/data/Jewellery_Sales_Report_2024_2026_Sorted.xlsx')
ws3 = wb3.active

random.seed(123)
payment_methods = ['Cash', 'Card', 'UPI', 'Bank Transfer', 'EMI']
cust_list = list(customer_ids.values())

sale_id = 1
sale_item_id = 1

for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, values_only=True):
    inv_date = row[0]
    prod_code = str(row[1]).strip() if row[1] else None
    purity = str(row[2]).strip() if row[2] else None
    net_wt = num(row[3])
    wastage = num(row[4])
    gold_rate = num(row[5])
    making_chr = num(row[6])
    sale_value = num(row[7])
    
    if not prod_code or sale_value <= 0:
        continue
    
    # Date handling
    if isinstance(inv_date, datetime):
        date_str = inv_date.strftime('%Y-%m-%d')
    else:
        date_str = str(inv_date).strip()[:10] if inv_date else '2025-01-01'
    
    # Get product_id
    prod_id = products.get(prod_code)
    if not prod_id:
        continue
    
    # Generate invoice number
    invoice = f"INV-{date_str.replace('-','')}-{sale_id:04d}"
    
    # Assign random customer
    cust_id = random.choice(cust_list)
    
    # Calculate tax
    subtotal = round(sale_value / 1.03, 2)  # Remove 3% GST
    tax_pct = 3.00
    tax_amt = round(sale_value - subtotal, 2)
    
    # Random payment method
    pay_method = random.choice(payment_methods)
    
    # Insert sale
    print(f"INSERT INTO sales (sale_id, customer_id, invoice_number, sale_date, subtotal, discount_percent, discount_amount, tax_percent, tax_amount, total_amount, payment_method, payment_status) VALUES")
    print(f"  ({sale_id}, {cust_id}, {esc(invoice)}, {esc(date_str)}, {subtotal}, 0.00, 0.00, {tax_pct}, {tax_amt}, {sale_value}, {esc(pay_method)}, 'Paid');")
    print()
    
    # Insert sale item
    print(f"INSERT INTO sale_items (sale_item_id, sale_id, product_id, quantity, unit_price, purity, net_weight, wastage_percent, gold_rate, making_charges, discount, total_price) VALUES")
    print(f"  ({sale_item_id}, {sale_id}, {prod_id}, 1, {sale_value}, {esc(purity)}, {net_wt}, {wastage}, {gold_rate}, {making_chr}, 0.00, {sale_value});")
    print()
    
    sale_id += 1
    sale_item_id += 1

print("-- ============================================================")
print(f"-- Seed complete: {len(categories)} categories, {len(products)} products,")
print(f"--   {len(inv_rows)} inventory records, {len(cust_rows)} customers, {sale_id-1} sales")
print("-- ============================================================")
print("SELECT 'Seed data imported successfully!' AS status;")
